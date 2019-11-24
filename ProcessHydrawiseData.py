"""
Read in .xlsx schedule data downloaded from Hydrawise.  Return it as a Schedule
"""
import numpy as np
import datetime
from openpyxl import load_workbook
from WaterModel import schedule

# spreadsheet may be empty.  In this case, there is a single sheet, 'Worksheet', and all cells are empty

def loadHydraData(inputFileName, controllerId, checkDate):
    wb = load_workbook(inputFileName)
    zoneList = wb.sheetnames

    sched = schedule(controllerId)

    for zone in zoneList:
        ws = wb[zone]
        if ws['A1'].value is None:  # check for empty
            continue
        
        assert(ws['A1'].value == 'Date')
        assert(ws['B1'].value == 'Time')
        assert(ws['C1'].value == 'min')
        
        for row in range(2, ws.max_row):
            cellBId = 'B{}'.format(row)
            cellCId = 'C{}'.format(row)
            zoneStart = ws[cellBId].value # datetime
            zoneDuration = ws[cellCId].value  # minutes
            if zoneStart is None:
                break
            if zoneStart.date() == checkDate.date():
                if zoneDuration > 0:
                    sched.addZone(zoneStart, zoneDuration, zone)
                    print('Hydrawise added: ', zoneStart.date(), zoneDuration, zone)
                break
            
    sched.finalize()

    return sched
        

    
